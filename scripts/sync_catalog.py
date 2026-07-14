#!/usr/bin/env python3
"""
Drive Exports Sync — reads Excel files from C:\agents\drive_exports\,
writes to SQLite dev.db AND generates MariaDB seed SQL.
"""
import openpyxl, json, datetime, time, os, sys, re, requests, sqlite3

DB_PATH = r"C:\Users\LokiF\Desktop\PDFWEBSITE\dev.db"
LOCAL = r"C:\agents\drive_exports"
OUTPUT_DIR = r"C:\Users\LokiF\Desktop\PDFWEBSITE\backend"
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "seed_drive_data.sql")
DATASET = "drive_data"


def esc(val):
    if val is None or val == "": return "NULL"
    if isinstance(val, (int, float)): return str(val)
    return "'" + str(val).replace("\\", "\\\\").replace("'", "''") + "'"

def parse_int(val):
    if val is None or str(val).strip() in ("", "-", "N/A", "n/a"): return None
    try: return int(str(val).replace(",", "").strip())
    except: return None

def parse_float(val):
    if val is None or str(val).strip() in ("", "-", "N/A", "n/a"): return None
    try: return float(str(val).replace("%", "").replace(",", "").strip())
    except: return None

def needs_refresh(db_path, dataset, new_update_date):
    try:
        conn = sqlite3.connect(db_path)
        c = conn.cursor()
        c.execute("SELECT MAX(update_date) FROM sync_log WHERE dataset = ?", (dataset,))
        row = c.fetchone()
        conn.close()
        if row and row[0] and str(row[0]) >= str(new_update_date):
            return False
    except Exception:
        pass
    return True

def record_sync(db_path, dataset, url, row_count, update_date):
    try:
        conn = sqlite3.connect(db_path)
        c = conn.cursor()
        c.execute("SELECT COUNT(*) FROM sync_log WHERE dataset = ?", (dataset,))
        exists = c.fetchone()[0]
        if exists:
            c.execute("UPDATE sync_log SET url=?, row_count=?, update_date=?, synced_at=? WHERE dataset=?",
                      (url, row_count, update_date, datetime.datetime.now().isoformat(), dataset))
        else:
            c.execute("INSERT INTO sync_log (dataset, url, row_count, update_date, synced_at) VALUES (?,?,?,?,?)",
                      (dataset, url, row_count, update_date, datetime.datetime.now().isoformat()))
        conn.commit()
        conn.close()
    except Exception as e:
        print(f"  sync_log warning: {e}")

def init_db(db_path):
    conn = sqlite3.connect(db_path)
    c = conn.cursor()
    c.execute("""CREATE TABLE IF NOT EXISTS sync_log (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        dataset TEXT NOT NULL,
        url TEXT,
        row_count INTEGER,
        update_date TEXT,
        synced_at TEXT
    )""")
    c.execute("""CREATE TABLE IF NOT EXISTS aggregate_catalog (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        cat_id TEXT,
        category TEXT,
        lane TEXT,
        source_type TEXT,
        evidence_note TEXT,
        scope_seen TEXT,
        result_use TEXT,
        confidence TEXT,
        source_ref TEXT
    )""")
    c.execute("""CREATE TABLE IF NOT EXISTS prr_tracker (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        request TEXT,
        matter_type TEXT,
        agency TEXT,
        stage TEXT,
        request_date TEXT,
        last_activity TEXT,
        deadline_regime TEXT,
        initial_response_due TEXT,
        initial_response_date TEXT,
        initial_response_timeliness TEXT,
        production_due TEXT,
        stated_agreed_due TEXT,
        deadline_basis TEXT,
        current_deadline_status TEXT,
        request_summary TEXT,
        timeframe_scope TEXT,
        custodian_scope TEXT,
        record_category_scope TEXT,
        search_terms TEXT,
        exclusions TEXT,
        responsive_records TEXT,
        missing_gaps TEXT,
        no_custody_claimed TEXT,
        withheld_exemptions_fee TEXT,
        scope_drift TEXT,
        appeal_determination TEXT,
        next_action TEXT,
        gmail_source TEXT,
        evidence_reviewed TEXT,
        confidence TEXT,
        notes TEXT
    )""")
    conn.commit()
    conn.close()

def read_excel_sheet(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=False)
    ws = wb.active
    headers = []
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=col).value
        headers.append(str(v).strip() if v else "")
    rows = []
    for r in range(2, ws.max_row + 1):
        vals = {}
        for col in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=col).value
            key = headers[col - 1]
            if v is not None and key:
                vals[key] = str(v).strip()
        if vals:
            rows.append(vals)
    wb.close()
    return rows

def write_aggregate_catalog_sqlite(db_path, rows):
    conn = sqlite3.connect(db_path)
    c = conn.cursor()
    c.execute("DELETE FROM aggregate_catalog")
    sql = """INSERT INTO aggregate_catalog
        (cat_id, category, lane, source_type, evidence_note, scope_seen, result_use, confidence, source_ref)
        VALUES (?,?,?,?,?,?,?,?,?)"""
    for r in rows:
        c.execute(sql, (
            r.get("ID"), r.get("Category"), r.get("Lane"),
            r.get("Source Type"), r.get("Evidence Note"), r.get("Scope Seen"),
            r.get("Result / Use"), r.get("Confidence"), r.get("Source")
        ))
    conn.commit()
    conn.close()

def write_prr_tracker_sqlite(db_path, rows):
    conn = sqlite3.connect(db_path)
    c = conn.cursor()
    c.execute("DELETE FROM prr_tracker")
    sql = """INSERT INTO prr_tracker
        (request, matter_type, agency, stage, request_date, last_activity, deadline_regime,
         initial_response_due, initial_response_date, initial_response_timeliness,
         production_due, stated_agreed_due, deadline_basis, current_deadline_status,
         request_summary, timeframe_scope, custodian_scope, record_category_scope,
         search_terms, exclusions, responsive_records, missing_gaps, no_custody_claimed,
         withheld_exemptions_fee, scope_drift, appeal_determination, next_action,
         gmail_source, evidence_reviewed, confidence, notes)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)"""
    for r in rows:
        c.execute(sql, (
            r.get("Request"), r.get("Matter Type"), r.get("Agency"),
            r.get("Stage"), r.get("Request Date"), r.get("Last Activity"),
            r.get("Applicable Deadline Regime(s)"), r.get("Initial Response Due"),
            r.get("Initial Response/Ack Date"), r.get("Initial Response Timeliness"),
            r.get("Standard Access/Production Due"), r.get("Stated/Agreed/Ordered Due"),
            r.get("Deadline Basis / Determination Terms"), r.get("Current Deadline Status"),
            r.get("Request Summary / Original Ask"), r.get("Timeframe Scope"),
            r.get("Custodian / Holder Scope"), r.get("Record Category / Subject Scope"),
            r.get("Search Terms / Named Entities"), r.get("Exclusions / Excluded Records"),
            r.get("Responsive Records Returned"), r.get("Missing/In-Scope Gaps"),
            r.get("No Custody / No Records Claimed"), r.get("Withheld/Redacted/Exemptions/Fee"),
            r.get("Out-of-Scope Produced / Scope Drift"), r.get("Appeal / Determination"),
            r.get("Next Action"), r.get("Gmail Source"),
            r.get("Attachment/Evidence Reviewed"), r.get("Confidence"), r.get("Notes")
        ))
    conn.commit()
    conn.close()


def main():
    print("=" * 60)
    print("Drive Exports Sync (C:\\agents\\drive_exports\\)")
    print("=" * 60)

    init_db(DB_PATH)
    today = datetime.date.today().isoformat()

    if not needs_refresh(DB_PATH, DATASET, today):
        print("Already synced today — skipping")
        return

    sql_lines = [
        "-- Parent Data Force v2 — Drive Data Sync",
        f"-- Generated {datetime.datetime.now().isoformat()}",
        "-- Tables: aggregate_catalog, prr_tracker",
        "",
    ]

    total_rows = 0
    xlsx_files = [f for f in os.listdir(LOCAL) if f.endswith(".xlsx")]
    if not xlsx_files:
        print("No .xlsx files found in drive_exports")
        return

    for fname in xlsx_files:
        path = os.path.join(LOCAL, fname)
        print(f"\nReading: {fname}")

        if "Aggregate" in fname:
            rows = read_excel_sheet(path)
            print(f"  Aggregate Catalog: {len(rows)} rows")
            write_aggregate_catalog_sqlite(DB_PATH, rows)
            total_rows += len(rows)

            sql_lines.append(f"-- Aggregate Catalog: {len(rows)} rows")
            sql_lines.append("DELETE FROM aggregate_catalog;")
            sql_lines.append("ALTER TABLE aggregate_catalog AUTO_INCREMENT = 1;")
            for r in rows:
                sql_lines.append(
                    f"INSERT INTO aggregate_catalog (cat_id, category, lane, source_type, evidence_note, scope_seen, result_use, confidence, source_ref) VALUES ("
                    f"{esc(r.get('ID'))}, {esc(r.get('Category'))}, {esc(r.get('Lane'))}, "
                    f"{esc(r.get('Source Type'))}, {esc(r.get('Evidence Note'))}, {esc(r.get('Scope Seen'))}, "
                    f"{esc(r.get('Result / Use'))}, {esc(r.get('Confidence'))}, {esc(r.get('Source'))});"
                )

        elif "Public Records" in fname:
            rows = read_excel_sheet(path)
            print(f"  PRR Tracker: {len(rows)} rows")
            write_prr_tracker_sqlite(DB_PATH, rows)
            total_rows += len(rows)

            sql_lines.append(f"\n-- PRR Tracker: {len(rows)} rows")
            sql_lines.append("DELETE FROM prr_tracker;")
            sql_lines.append("ALTER TABLE prr_tracker AUTO_INCREMENT = 1;")
            for r in rows:
                sql_lines.append(
                    f"INSERT INTO prr_tracker (request, matter_type, agency, stage, request_date, last_activity, deadline_regime, "
                    f"initial_response_due, initial_response_date, initial_response_timeliness, production_due, stated_agreed_due, "
                    f"deadline_basis, current_deadline_status, request_summary, timeframe_scope, custodian_scope, record_category_scope, "
                    f"search_terms, exclusions, responsive_records, missing_gaps, no_custody_claimed, withheld_exemptions_fee, "
                    f"scope_drift, appeal_determination, next_action, gmail_source, evidence_reviewed, confidence, notes) VALUES ("
                    f"{esc(r.get('Request'))}, {esc(r.get('Matter Type'))}, {esc(r.get('Agency'))}, "
                    f"{esc(r.get('Stage'))}, {esc(r.get('Request Date'))}, {esc(r.get('Last Activity'))}, "
                    f"{esc(r.get('Applicable Deadline Regime(s)'))}, {esc(r.get('Initial Response Due'))}, "
                    f"{esc(r.get('Initial Response/Ack Date'))}, {esc(r.get('Initial Response Timeliness'))}, "
                    f"{esc(r.get('Standard Access/Production Due'))}, {esc(r.get('Stated/Agreed/Ordered Due'))}, "
                    f"{esc(r.get('Deadline Basis / Determination Terms'))}, {esc(r.get('Current Deadline Status'))}, "
                    f"{esc(r.get('Request Summary / Original Ask'))}, {esc(r.get('Timeframe Scope'))}, "
                    f"{esc(r.get('Custodian / Holder Scope'))}, {esc(r.get('Record Category / Subject Scope'))}, "
                    f"{esc(r.get('Search Terms / Named Entities'))}, {esc(r.get('Exclusions / Excluded Records'))}, "
                    f"{esc(r.get('Responsive Records Returned'))}, {esc(r.get('Missing/In-Scope Gaps'))}, "
                    f"{esc(r.get('No Custody / No Records Claimed'))}, {esc(r.get('Withheld/Redacted/Exemptions/Fee'))}, "
                    f"{esc(r.get('Out-of-Scope Produced / Scope Drift'))}, {esc(r.get('Appeal / Determination'))}, "
                    f"{esc(r.get('Next Action'))}, {esc(r.get('Gmail Source'))}, "
                    f"{esc(r.get('Attachment/Evidence Reviewed'))}, {esc(r.get('Confidence'))}, {esc(r.get('Notes'))});"
                )
        else:
            print(f"  Unknown file type — skipping")

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
        f.write("\n".join(sql_lines))

    record_sync(DB_PATH, DATASET, LOCAL, total_rows, today)

    print(f"\nSQL written: {OUTPUT_FILE} ({os.path.getsize(OUTPUT_FILE):,} bytes)")
    print(f"SQLite: {total_rows} rows across aggregate_catalog + prr_tracker")
    print("Done.")
    print()

if __name__ == "__main__":
    main()
