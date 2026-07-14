#!/usr/bin/env python3
"""
Parent Data Force — Google Drive Data Sync
Reads exported Excel files from G:\My Drive\* and generates SQL seed files.
"""
import os, json, datetime, shutil
from pathlib import Path

try:
    import openpyxl
except ImportError:
    import subprocess, sys
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl

ROOT = Path(__file__).resolve().parent.parent
DRIVE_XLSX_DIR = r"G:\My Drive\Parent Data Force XLSX Exports - 2026-07-10"
LOCAL_CACHE = ROOT / "drive_exports"
OUTPUT_SQL  = ROOT / "backend" / "seed_drive_data.sql"

# Also check older PRS data
PRS_XLSX = r"G:\My Drive\Prs\PRS_Intakes_Received_and_Findings_since_1.1.21 (1).xlsx"

def escape(val):
    if val is None: return "NULL"
    s = str(val).replace("\\", "\\\\").replace("'", "''")
    return f"'{s}'"

def read_excel(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    headers = []
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=col).value
        headers.append(str(v).strip() if v else "")
    rows = []
    for row in range(2, ws.max_row + 1):
        r = {}
        for col in range(1, ws.max_column + 1):
            v = ws.cell(row=row, column=col).value
            key = headers[col - 1]
            if v is not None and key:
                r[key] = str(v).strip()
        if r:
            rows.append(r)
    return rows

def main():
    print("=" * 60)
    print("Parent Data Force — Google Drive Data Sync")
    print("=" * 60)

    sql_lines = [
        "-- Parent Data Force v2 — Google Drive Data Sync",
        f"-- Generated {datetime.datetime.now().isoformat()}",
        "-- Tables: aggregate_catalog, prr_tracker, prs_data",
        "",
        "CREATE TABLE IF NOT EXISTS aggregate_catalog (",
        "    id INTEGER PRIMARY KEY AUTOINCREMENT,",
        "    cat_id TEXT,",
        "    category TEXT,",
        "    lane TEXT,",
        "    source_type TEXT,",
        "    evidence_note TEXT,",
        "    scope_seen TEXT,",
        "    result_use TEXT,",
        "    confidence TEXT,",
        "    source_ref TEXT",
        ");",
        "",
        "CREATE TABLE IF NOT EXISTS prr_tracker (",
        "    id INTEGER PRIMARY KEY AUTOINCREMENT,",
        "    request TEXT,",
        "    matter_type TEXT,",
        "    agency TEXT,",
        "    stage TEXT,",
        "    request_date TEXT,",
        "    last_activity TEXT,",
        "    deadline_regime TEXT,",
        "    initial_response_due TEXT,",
        "    initial_response_date TEXT,",
        "    initial_response_timeliness TEXT,",
        "    production_due TEXT,",
        "    stated_agreed_due TEXT,",
        "    deadline_basis TEXT,",
        "    current_deadline_status TEXT,",
        "    request_summary TEXT,",
        "    timeframe_scope TEXT,",
        "    custodian_scope TEXT,",
        "    record_category_scope TEXT,",
        "    search_terms TEXT,",
        "    exclusions TEXT,",
        "    responsive_records TEXT,",
        "    missing_gaps TEXT,",
        "    no_custody_claimed TEXT,",
        "    withheld_exemptions_fee TEXT,",
        "    scope_drift TEXT,",
        "    appeal_determination TEXT,",
        "    next_action TEXT,",
        "    gmail_source TEXT,",
        "    evidence_reviewed TEXT,",
        "    confidence TEXT,",
        "    notes TEXT",
        ");",
        "",
        "CREATE TABLE IF NOT EXISTS prs_data (",
        "    prs_number TEXT PRIMARY KEY,",
        "    district TEXT,",
        "    intake_date TEXT,",
        "    status TEXT,",
        "    findings_date TEXT,",
        "    category TEXT,",
        "    subcategory TEXT,",
        "    closure_code TEXT",
        ");",
        "",
    ]

    # 1. Aggregate Request Catalog
    cat_file = None
    if os.path.isdir(DRIVE_XLSX_DIR):
        for f in os.listdir(DRIVE_XLSX_DIR):
            if "Aggregate" in f and f.endswith(".xlsx"):
                cat_file = os.path.join(DRIVE_XLSX_DIR, f)
                break

    if cat_file and os.path.exists(cat_file):
        print(f"Reading Aggregate Catalog: {cat_file}")
        rows = read_excel(cat_file)
        sql_lines.append(f"-- Aggregate Catalog: {len(rows)} rows")
        for r in rows:
            sql_lines.append(
                f"INSERT INTO aggregate_catalog (cat_id, category, lane, source_type, evidence_note, scope_seen, result_use, confidence, source_ref) VALUES ("
                f"{escape(r.get('ID'))}, {escape(r.get('Category'))}, {escape(r.get('Lane'))}, "
                f"{escape(r.get('Source Type'))}, {escape(r.get('Evidence Note'))}, {escape(r.get('Scope Seen'))}, "
                f"{escape(r.get('Result / Use'))}, {escape(r.get('Confidence'))}, {escape(r.get('Source'))});"
            )
        print(f"  {len(rows)} catalog entries")
    else:
        print("Aggregate Catalog not found on G: drive — skipping")

    # 2. PRR Tracker
    prr_file = None
    if os.path.isdir(DRIVE_XLSX_DIR):
        for f in os.listdir(DRIVE_XLSX_DIR):
            if "Public Records Request Tracker" in f and f.endswith(".xlsx"):
                prr_file = os.path.join(DRIVE_XLSX_DIR, f)
                break

    if prr_file and os.path.exists(prr_file):
        print(f"Reading PRR Tracker: {prr_file}")
        rows = read_excel(prr_file)
        sql_lines.append(f"\n-- PRR Tracker: {len(rows)} rows")
        for r in rows:
            sql_lines.append(
                f"INSERT INTO prr_tracker (request, matter_type, agency, stage, request_date, last_activity, deadline_regime, "
                f"initial_response_due, initial_response_date, initial_response_timeliness, production_due, stated_agreed_due, "
                f"deadline_basis, current_deadline_status, request_summary, timeframe_scope, custodian_scope, record_category_scope, "
                f"search_terms, exclusions, responsive_records, missing_gaps, no_custody_claimed, withheld_exemptions_fee, "
                f"scope_drift, appeal_determination, next_action, gmail_source, evidence_reviewed, confidence, notes) VALUES ("
                f"{escape(r.get('Request'))}, {escape(r.get('Matter Type'))}, {escape(r.get('Agency'))}, "
                f"{escape(r.get('Stage'))}, {escape(r.get('Request Date'))}, {escape(r.get('Last Activity'))}, "
                f"{escape(r.get('Applicable Deadline Regime(s)'))}, {escape(r.get('Initial Response Due'))}, "
                f"{escape(r.get('Initial Response/Ack Date'))}, {escape(r.get('Initial Response Timeliness'))}, "
                f"{escape(r.get('Standard Access/Production Due'))}, {escape(r.get('Stated/Agreed/Ordered Due'))}, "
                f"{escape(r.get('Deadline Basis / Determination Terms'))}, {escape(r.get('Current Deadline Status'))}, "
                f"{escape(r.get('Request Summary / Original Ask'))}, {escape(r.get('Timeframe Scope'))}, "
                f"{escape(r.get('Custodian / Holder Scope'))}, {escape(r.get('Record Category / Subject Scope'))}, "
                f"{escape(r.get('Search Terms / Named Entities'))}, {escape(r.get('Exclusions / Excluded Records'))}, "
                f"{escape(r.get('Responsive Records Returned'))}, {escape(r.get('Missing/In-Scope Gaps'))}, "
                f"{escape(r.get('No Custody / No Records Claimed'))}, {escape(r.get('Withheld/Redacted/Exemptions/Fee'))}, "
                f"{escape(r.get('Out-of-Scope Produced / Scope Drift'))}, {escape(r.get('Appeal / Determination'))}, "
                f"{escape(r.get('Next Action'))}, {escape(r.get('Gmail Source'))}, "
                f"{escape(r.get('Attachment/Evidence Reviewed'))}, {escape(r.get('Confidence'))}, {escape(r.get('Notes'))});"
            )
        print(f"  {len(rows)} tracker entries")
    else:
        print("PRR Tracker not found on G: drive — skipping")

    # 3. PRS Data
    if os.path.exists(PRS_XLSX):
        print(f"Reading PRS Data: {PRS_XLSX}")
        rows = read_excel(PRS_XLSX)
        sql_lines.append(f"\n-- PRS Data: {len(rows)} rows")
        for r in rows:
            sql_lines.append(
                f"INSERT OR IGNORE INTO prs_data (prs_number, district, intake_date, status, findings_date, category, subcategory, closure_code) VALUES ("
                f"{escape(r.get('Number'))}, {escape(r.get('District / Agency'))}, {escape(r.get('Intake Form Received'))}, "
                f"{escape(r.get('Status'))}, {escape(r.get('Letter of Findings Issued'))}, "
                f"{escape(r.get('Category'))}, {escape(r.get('Subcategory'))}, {escape(r.get('Closure Code'))});"
            )
        print(f"  {len(rows)} PRS entries (12,464 total)")
    else:
        print("PRS Data not found on G: drive — skipping")

    OUTPUT_SQL.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_SQL.write_text("\n".join(sql_lines), encoding="utf-8")
    print(f"\nSQL written: {OUTPUT_SQL} ({OUTPUT_SQL.stat().st_size:,} bytes)")
    print("Done.")

if __name__ == "__main__":
    main()
