from __future__ import annotations

from collections import defaultdict
from email import policy
from email.header import decode_header, make_header
from email.parser import BytesParser
import hashlib
from io import BytesIO
from pathlib import Path
import re

import openpyxl

from .common import ensure_dir, slugify, write_json


SUMMARY_TOTAL_CODE = "99999997"
SPECIAL_PROGRAM_KEYWORDS = (
    "therapeutic",
    "public day",
    "day school",
    "assessment center",
    "specialized",
    "special education",
    "special ed",
    "collaborative",
    "exceptional",
    "alternative",
)


def rel_path(path: Path, root: Path) -> str:
    return str(path.relative_to(root)).replace("\\", "/")


def sha256_bytes(payload: bytes) -> str:
    return hashlib.sha256(payload).hexdigest()


def sha256_path(path: Path) -> str:
    return sha256_bytes(path.read_bytes())


def clean_code(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return text.zfill(8) if text.isdigit() and len(text) <= 8 else text


def decode_rfc2047(value: str) -> str:
    try:
        return str(make_header(decode_header(value)))
    except Exception:
        return value


def parse_year_label(value: str) -> str:
    match = re.search(r"(20\d{2}-\d{2})", value)
    return match.group(1) if match else "unknown-year"


def canonical_workbook_name(year_label: str) -> str:
    return f"dese-student-restraints-{year_label}-public-schools.xlsx"


def to_int(value: object) -> int:
    if value is None:
        return 0
    try:
        return int(float(str(value).strip()))
    except (TypeError, ValueError):
        return 0


def ratio(a: float | int | None, b: float | int | None) -> float | None:
    if a is None or b in (None, 0):
        return None
    return round(float(a) / float(b), 4)


def rate_per_100(numerator: float | int | None, enrollment: float | int | None) -> float | None:
    if numerator is None or enrollment in (None, 0):
        return None
    return round((float(numerator) / float(enrollment)) * 100.0, 4)


def percentile_linear(values: list[float], q: float) -> float | None:
    if not values:
        return None
    if len(values) == 1:
        return round(values[0], 4)
    idx = (len(values) - 1) * q
    lo = int(idx)
    hi = min(lo + 1, len(values) - 1)
    weight = idx - lo
    result = values[lo] * (1.0 - weight) + values[hi] * weight
    return round(result, 4)


def parse_metric(value: object) -> dict:
    if value is None:
        return {"reported": 0, "suppressed": False, "raw": 0}

    text = str(value).strip()
    if text == "":
        return {"reported": 0, "suppressed": False, "raw": 0}
    if text == "-":
        return {"reported": "suppressed", "suppressed": True, "raw": None}

    try:
        numeric = int(float(text))
        return {"reported": numeric, "suppressed": False, "raw": numeric}
    except (TypeError, ValueError):
        return {"reported": None, "suppressed": False, "raw": None}


def is_summary_row(district_code: str, school_code: str, district_name: str, school_name: str) -> bool:
    if district_code == SUMMARY_TOTAL_CODE or school_code == SUMMARY_TOTAL_CODE:
        return True
    combined = f"{district_name} {school_name}".strip().lower()
    return "public schools total" in combined


def traditional_lens_status(school_name: str, enrollment: int, summary_row: bool) -> tuple[bool, str | None]:
    if summary_row:
        return False, "summary-row"
    if enrollment < 100:
        return False, "enrollment-under-100"
    lowered = school_name.lower()
    for keyword in SPECIAL_PROGRAM_KEYWORDS:
        if keyword in lowered:
            return False, f"keyword:{keyword}"
    return True, None


def workbook_year_from_bytes(payload: bytes) -> str:
    workbook = openpyxl.load_workbook(BytesIO(payload), read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    first_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
    year_label = parse_year_label(str(first_row[0] or ""))
    workbook.close()
    return year_label


def write_binary_if_changed(path: Path, payload: bytes) -> bool:
    ensure_dir(path.parent)
    if path.exists() and sha256_path(path) == sha256_bytes(payload):
        return False
    path.write_bytes(payload)
    return True


def extract_eml_attachments(root: Path) -> dict:
    intake_dir = root / "intake"
    raw_dir = root / "data" / "raw" / "dese-restraint"
    note_dir = raw_dir / "reference"
    ensure_dir(raw_dir)
    ensure_dir(note_dir)

    workbook_targets: set[Path] = set()
    workbook_records: list[dict] = []
    reference_docs: list[dict] = []

    for eml_path in sorted(intake_dir.glob("*.eml")):
        message = BytesParser(policy=policy.default).parsebytes(eml_path.read_bytes())
        for part in message.walk():
            filename = part.get_filename()
            if not filename:
                continue
            decoded_name = decode_rfc2047(filename)
            payload = part.get_payload(decode=True) or b""
            lowered = decoded_name.lower()

            if lowered.endswith(".xlsx"):
                year_label = workbook_year_from_bytes(payload)
                if year_label == "unknown-year":
                    continue
                target = raw_dir / canonical_workbook_name(year_label)
                changed = write_binary_if_changed(target, payload)
                workbook_targets.add(target)
                workbook_records.append(
                    {
                        "sourceType": "eml-attachment",
                        "sourceEml": rel_path(eml_path, root),
                        "attachmentName": decoded_name,
                        "schoolYear": year_label,
                        "target": rel_path(target, root),
                        "sha256": sha256_bytes(payload),
                        "updated": changed,
                    }
                )

            if lowered.endswith(".docx"):
                safe_name = f"{slugify(Path(decoded_name).stem)}.docx"
                target = note_dir / safe_name
                changed = write_binary_if_changed(target, payload)
                reference_docs.append(
                    {
                        "sourceType": "eml-attachment",
                        "sourceEml": rel_path(eml_path, root),
                        "attachmentName": decoded_name,
                        "target": rel_path(target, root),
                        "sha256": sha256_bytes(payload),
                        "updated": changed,
                    }
                )

    return {
        "workbookTargets": sorted(workbook_targets),
        "workbookRecords": workbook_records,
        "referenceDocs": reference_docs,
    }


def ingest_loose_intake_workbooks(root: Path) -> dict:
    intake_dir = root / "intake"
    raw_dir = root / "data" / "raw" / "dese-restraint"
    ensure_dir(raw_dir)

    workbook_targets: set[Path] = set()
    workbook_records: list[dict] = []

    for source in sorted(intake_dir.glob("restraints*.xlsx")):
        payload = source.read_bytes()
        year_label = workbook_year_from_bytes(payload)
        if year_label == "unknown-year":
            continue
        target = raw_dir / canonical_workbook_name(year_label)
        changed = write_binary_if_changed(target, payload)
        workbook_targets.add(target)
        workbook_records.append(
            {
                "sourceType": "loose-intake-xlsx",
                "sourceFile": rel_path(source, root),
                "schoolYear": year_label,
                "target": rel_path(target, root),
                "sha256": sha256_bytes(payload),
                "updated": changed,
            }
        )

    return {"workbookTargets": sorted(workbook_targets), "workbookRecords": workbook_records}


def select_workbooks(root: Path) -> list[Path]:
    raw_dir = root / "data" / "raw" / "dese-restraint"
    ensure_dir(raw_dir)

    by_year: dict[str, list[Path]] = defaultdict(list)
    for candidate in sorted(raw_dir.glob("*.xlsx")):
        year_label = parse_year_label(candidate.name)
        if year_label == "unknown-year":
            try:
                year_label = workbook_year_from_bytes(candidate.read_bytes())
            except Exception:
                year_label = "unknown-year"
        if year_label == "unknown-year":
            continue
        by_year[year_label].append(candidate)

    selected: list[Path] = []
    for year_label, candidates in by_year.items():
        preferred_name = canonical_workbook_name(year_label)
        preferred = next((path for path in candidates if path.name == preferred_name), None)
        if preferred is not None:
            selected.append(preferred)
            continue
        selected.append(sorted(candidates, key=lambda p: p.stat().st_mtime, reverse=True)[0])

    selected.sort(key=lambda p: parse_year_label(p.name))
    return selected


def build_school_year_records(workbooks: list[Path]) -> list[dict]:
    records: list[dict] = []

    for workbook in workbooks:
        wb = openpyxl.load_workbook(workbook, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        title_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        school_year = parse_year_label(str(title_row[0] or ""))

        for row in ws.iter_rows(min_row=3, values_only=True):
            if not row or row[0] is None:
                continue

            district_code = clean_code(row[0])
            district_name = str(row[1] or "").strip()
            school_code = clean_code(row[2])
            school_name = str(row[3] or "").strip()
            enrollment = to_int(row[4])

            students = parse_metric(row[5])
            restraints = parse_metric(row[6])
            injuries = parse_metric(row[7])

            summary_row = is_summary_row(district_code, school_code, district_name, school_name)
            traditional_ok, traditional_reason = traditional_lens_status(school_name, enrollment, summary_row)

            raw_students = students["raw"]
            raw_restraints = restraints["raw"]
            raw_injuries = injuries["raw"]

            records.append(
                {
                    "id": slugify(f"{school_year}-{district_code}-{school_code}"),
                    "schoolYear": school_year,
                    "districtCode": district_code,
                    "districtName": district_name,
                    "schoolCode": school_code,
                    "schoolName": school_name,
                    "schoolKey": f"{district_code}::{school_code}",
                    "enrollment": enrollment,
                    "studentsRestrained": {
                        "reported": students["reported"],
                        "suppressed": students["suppressed"],
                    },
                    "totalRestraints": {
                        "reported": restraints["reported"],
                        "suppressed": restraints["suppressed"],
                    },
                    "totalInjuries": {
                        "reported": injuries["reported"],
                        "suppressed": injuries["suppressed"],
                    },
                    "rawStudentsRestrained": raw_students,
                    "rawTotalRestraints": raw_restraints,
                    "rawTotalInjuries": raw_injuries,
                    "rawStudentsRestrainedRatePer100": rate_per_100(raw_students, enrollment),
                    "rawRestraintRatePer100": rate_per_100(raw_restraints, enrollment),
                    "rawInjuryRatePer100": rate_per_100(raw_injuries, enrollment),
                    "rawRestraintsPerStudentRestrained": ratio(raw_restraints, raw_students),
                    "rawInjuriesPerRestraint": ratio(raw_injuries, raw_restraints),
                    "isRawKnownRestraints": raw_restraints is not None,
                    "isRawKnownStudents": raw_students is not None,
                    "isRawKnownInjuries": raw_injuries is not None,
                    "isSuppressedAny": students["suppressed"] or restraints["suppressed"] or injuries["suppressed"],
                    "isSummaryRow": summary_row,
                    "isTraditionalComparable": traditional_ok,
                    "traditionalExclusionReason": traditional_reason,
                    "source": "DESE Student Restraints Report - Public Schools",
                    "sourceWorkbook": workbook.name,
                }
            )

        wb.close()

    records.sort(key=lambda row: (row.get("schoolYear", ""), row.get("districtCode", ""), row.get("schoolCode", "")))
    return records


def add_school_year_rankings(records: list[dict]) -> None:
    by_year: dict[str, list[dict]] = defaultdict(list)
    for row in records:
        by_year[row.get("schoolYear", "")].append(row)

    for rows in by_year.values():
        all_visible = [
            row
            for row in rows
            if not row.get("isSummaryRow") and row.get("isRawKnownRestraints") and row.get("rawRestraintRatePer100") is not None
        ]
        trad_visible = [row for row in all_visible if row.get("isTraditionalComparable")]

        all_by_total = sorted(all_visible, key=lambda row: (-(row.get("rawTotalRestraints") or 0), row.get("id", "")))
        all_by_rate = sorted(all_visible, key=lambda row: (-(row.get("rawRestraintRatePer100") or 0.0), row.get("id", "")))
        trad_by_total = sorted(trad_visible, key=lambda row: (-(row.get("rawTotalRestraints") or 0), row.get("id", "")))
        trad_by_rate = sorted(trad_visible, key=lambda row: (-(row.get("rawRestraintRatePer100") or 0.0), row.get("id", "")))

        all_total_rank = {row.get("id", ""): idx + 1 for idx, row in enumerate(all_by_total)}
        all_rate_rank = {row.get("id", ""): idx + 1 for idx, row in enumerate(all_by_rate)}
        trad_total_rank = {row.get("id", ""): idx + 1 for idx, row in enumerate(trad_by_total)}
        trad_rate_rank = {row.get("id", ""): idx + 1 for idx, row in enumerate(trad_by_rate)}

        all_rates = sorted([row.get("rawRestraintRatePer100") for row in all_visible if row.get("rawRestraintRatePer100") is not None])
        trad_rates = sorted([row.get("rawRestraintRatePer100") for row in trad_visible if row.get("rawRestraintRatePer100") is not None])

        all_context = {
            "schoolCount": len(all_visible),
            "medianRatePer100": percentile_linear(all_rates, 0.5),
            "p95RatePer100": percentile_linear(all_rates, 0.95),
        }
        trad_context = {
            "schoolCount": len(trad_visible),
            "medianRatePer100": percentile_linear(trad_rates, 0.5),
            "p95RatePer100": percentile_linear(trad_rates, 0.95),
        }

        for row in rows:
            row_id = row.get("id", "")
            row["statewideContextAllRows"] = all_context
            row["statewideRankAllRows"] = {
                "byTotalRestraints": all_total_rank.get(row_id),
                "byRatePer100": all_rate_rank.get(row_id),
            }
            row["statewideContextTraditional"] = trad_context
            row["statewideRankTraditional"] = {
                "byTotalRestraints": trad_total_rank.get(row_id),
                "byRatePer100": trad_rate_rank.get(row_id),
            }


def build_district_year_metrics(records: list[dict], use_traditional_lens: bool) -> dict[str, dict[str, dict]]:
    grouped: dict[str, dict[str, list[dict]]] = defaultdict(lambda: defaultdict(list))
    for row in records:
        if row.get("isSummaryRow"):
            continue
        if use_traditional_lens and not row.get("isTraditionalComparable"):
            continue
        grouped[row.get("schoolYear", "")][row.get("districtCode", "")].append(row)

    result: dict[str, dict[str, dict]] = defaultdict(dict)
    for year, district_rows in grouped.items():
        for district_code, rows in district_rows.items():
            known_restraint_rows = [row for row in rows if row.get("isRawKnownRestraints")]
            known_students_rows = [row for row in rows if row.get("isRawKnownStudents")]
            known_injury_rows = [row for row in rows if row.get("isRawKnownInjuries")]

            total_restraints = sum(row.get("rawTotalRestraints") or 0 for row in known_restraint_rows)
            total_students = sum(row.get("rawStudentsRestrained") or 0 for row in known_students_rows)
            total_injuries = sum(row.get("rawTotalInjuries") or 0 for row in known_injury_rows)

            restraint_enrollment = sum(row.get("enrollment") or 0 for row in known_restraint_rows)
            student_enrollment = sum(row.get("enrollment") or 0 for row in known_students_rows)
            injury_enrollment = sum(row.get("enrollment") or 0 for row in known_injury_rows)

            paired_students_rows = [
                row for row in rows if row.get("isRawKnownRestraints") and row.get("isRawKnownStudents")
            ]
            paired_injury_rows = [
                row for row in rows if row.get("isRawKnownRestraints") and row.get("isRawKnownInjuries")
            ]

            paired_restraints_students = sum(row.get("rawTotalRestraints") or 0 for row in paired_students_rows)
            paired_students_total = sum(row.get("rawStudentsRestrained") or 0 for row in paired_students_rows)
            paired_restraints_injuries = sum(row.get("rawTotalRestraints") or 0 for row in paired_injury_rows)
            paired_injuries_total = sum(row.get("rawTotalInjuries") or 0 for row in paired_injury_rows)

            top_school = None
            if known_restraint_rows:
                winner = sorted(
                    known_restraint_rows,
                    key=lambda row: (
                        -(row.get("rawTotalRestraints") or 0),
                        -(row.get("rawRestraintRatePer100") or 0.0),
                        row.get("schoolName", ""),
                    ),
                )[0]
                top_school = {
                    "schoolKey": winner.get("schoolKey"),
                    "schoolCode": winner.get("schoolCode"),
                    "schoolName": winner.get("schoolName"),
                    "rawTotalRestraints": winner.get("rawTotalRestraints"),
                    "rawRestraintRatePer100": winner.get("rawRestraintRatePer100"),
                }

            result[year][district_code] = {
                "districtCode": district_code,
                "districtName": rows[0].get("districtName", district_code),
                "schoolYear": year,
                "schoolsInYear": len(rows),
                "knownRestraintRows": len(known_restraint_rows),
                "rawTotals": {
                    "restraints": total_restraints,
                    "studentsRestrained": total_students,
                    "injuries": total_injuries,
                },
                "rawRestraintRatePer100": rate_per_100(total_restraints, restraint_enrollment)
                if known_restraint_rows
                else None,
                "rawStudentsRestrainedRatePer100": rate_per_100(total_students, student_enrollment)
                if known_students_rows
                else None,
                "rawInjuryRatePer100": rate_per_100(total_injuries, injury_enrollment)
                if known_injury_rows
                else None,
                "restraintsPerStudentRestrainedRaw": ratio(paired_restraints_students, paired_students_total),
                "injuriesPerRestraintRaw": ratio(paired_injuries_total, paired_restraints_injuries),
                "topSchool": top_school,
            }

    return result


def apply_district_year_rankings(metrics_by_year: dict[str, dict[str, dict]]) -> None:
    for district_metrics in metrics_by_year.values():
        visible = [m for m in district_metrics.values() if m.get("rawRestraintRatePer100") is not None]
        by_total = sorted(visible, key=lambda item: (-(item.get("rawTotals", {}).get("restraints") or 0), item.get("districtCode", "")))
        by_rate = sorted(visible, key=lambda item: (-(item.get("rawRestraintRatePer100") or 0.0), item.get("districtCode", "")))

        rank_total = {m.get("districtCode", ""): idx + 1 for idx, m in enumerate(by_total)}
        rank_rate = {m.get("districtCode", ""): idx + 1 for idx, m in enumerate(by_rate)}

        rates = sorted([m.get("rawRestraintRatePer100") for m in visible if m.get("rawRestraintRatePer100") is not None])
        context = {
            "districtCount": len(visible),
            "medianRatePer100": percentile_linear(rates, 0.5),
            "p95RatePer100": percentile_linear(rates, 0.95),
        }

        for metric in district_metrics.values():
            code = metric.get("districtCode", "")
            metric["statewideContext"] = context
            metric["statewideRank"] = {
                "byTotalRestraints": rank_total.get(code),
                "byRatePer100": rank_rate.get(code),
            }


def build_school_rollup(records: list[dict]) -> list[dict]:
    grouped: dict[str, list[dict]] = defaultdict(list)
    for row in records:
        if row.get("isSummaryRow"):
            continue
        grouped[row.get("schoolKey", "")].append(row)

    rollup: list[dict] = []
    for school_key, rows in grouped.items():
        rows_sorted = sorted(rows, key=lambda row: row.get("schoolYear", ""))
        visible_rows = [row for row in rows_sorted if row.get("isRawKnownRestraints")]

        latest = visible_rows[-1] if visible_rows else None
        previous = visible_rows[-2] if len(visible_rows) > 1 else None

        peak_total = max(visible_rows, key=lambda row: row.get("rawTotalRestraints") or 0, default=None)
        peak_rate = max(visible_rows, key=lambda row: row.get("rawRestraintRatePer100") or 0.0, default=None)

        known_students_rows = [row for row in rows_sorted if row.get("isRawKnownStudents")]
        known_injury_rows = [row for row in rows_sorted if row.get("isRawKnownInjuries")]

        known_total_restraints = sum(row.get("rawTotalRestraints") or 0 for row in visible_rows)
        known_total_students = sum(row.get("rawStudentsRestrained") or 0 for row in known_students_rows)
        known_total_injuries = sum(row.get("rawTotalInjuries") or 0 for row in known_injury_rows)
        known_enrollment = sum(row.get("enrollment") or 0 for row in visible_rows)

        latest_all_rank = latest.get("statewideRankAllRows", {}) if latest else {}
        latest_all_context = latest.get("statewideContextAllRows", {}) if latest else {}
        latest_trad_rank = latest.get("statewideRankTraditional", {}) if latest else {}
        latest_trad_context = latest.get("statewideContextTraditional", {}) if latest else {}

        rollup.append(
            {
                "schoolKey": school_key,
                "districtCode": rows_sorted[0].get("districtCode", ""),
                "districtName": rows_sorted[0].get("districtName", ""),
                "schoolCode": rows_sorted[0].get("schoolCode", ""),
                "schoolName": rows_sorted[0].get("schoolName", ""),
                "isTraditionalComparable": rows_sorted[0].get("isTraditionalComparable", False),
                "traditionalExclusionReason": rows_sorted[0].get("traditionalExclusionReason"),
                "years": [row.get("schoolYear") for row in rows_sorted],
                "rowCount": len(rows_sorted),
                "visibleYearCount": len(visible_rows),
                "latestYear": latest.get("schoolYear") if latest else None,
                "latestMetrics": {
                    "rawTotalRestraints": latest.get("rawTotalRestraints") if latest else None,
                    "rawStudentsRestrained": latest.get("rawStudentsRestrained") if latest else None,
                    "rawTotalInjuries": latest.get("rawTotalInjuries") if latest else None,
                    "enrollment": latest.get("enrollment") if latest else None,
                    "rawRestraintRatePer100": latest.get("rawRestraintRatePer100") if latest else None,
                    "rawStudentsRestrainedRatePer100": latest.get("rawStudentsRestrainedRatePer100") if latest else None,
                    "rawInjuryRatePer100": latest.get("rawInjuryRatePer100") if latest else None,
                    "restraintsPerStudentRestrainedRaw": latest.get("rawRestraintsPerStudentRestrained") if latest else None,
                    "injuriesPerRestraintRaw": latest.get("rawInjuriesPerRestraint") if latest else None,
                },
                "previousVisibleYear": previous.get("schoolYear") if previous else None,
                "yearOverYear": {
                    "deltaRestraints": (latest.get("rawTotalRestraints") or 0) - (previous.get("rawTotalRestraints") or 0)
                    if latest and previous
                    else None,
                    "deltaRatePer100": round((latest.get("rawRestraintRatePer100") or 0.0) - (previous.get("rawRestraintRatePer100") or 0.0), 4)
                    if latest and previous and latest.get("rawRestraintRatePer100") is not None and previous.get("rawRestraintRatePer100") is not None
                    else None,
                    "pctChangeRestraints": ratio(
                        ((latest.get("rawTotalRestraints") or 0) - (previous.get("rawTotalRestraints") or 0)) * 100.0,
                        previous.get("rawTotalRestraints") if previous else None,
                    )
                    if latest and previous
                    else None,
                },
                "peakMetrics": {
                    "peakTotalYear": peak_total.get("schoolYear") if peak_total else None,
                    "peakTotalRestraints": peak_total.get("rawTotalRestraints") if peak_total else None,
                    "peakRateYear": peak_rate.get("schoolYear") if peak_rate else None,
                    "peakRatePer100": peak_rate.get("rawRestraintRatePer100") if peak_rate else None,
                },
                "cumulativeKnown": {
                    "restraints": known_total_restraints,
                    "studentsRestrained": known_total_students,
                    "injuries": known_total_injuries,
                    "enrollment": known_enrollment,
                    "weightedRestraintRatePer100": rate_per_100(known_total_restraints, known_enrollment),
                },
                "statewideLatestAllRows": {
                    "schoolYear": latest.get("schoolYear") if latest else None,
                    "rankByTotalRestraints": latest_all_rank.get("byTotalRestraints"),
                    "rankByRatePer100": latest_all_rank.get("byRatePer100"),
                    "schoolCount": latest_all_context.get("schoolCount"),
                    "medianRatePer100": latest_all_context.get("medianRatePer100"),
                    "p95RatePer100": latest_all_context.get("p95RatePer100"),
                },
                "statewideLatestTraditional": {
                    "schoolYear": latest.get("schoolYear") if latest else None,
                    "rankByTotalRestraints": latest_trad_rank.get("byTotalRestraints"),
                    "rankByRatePer100": latest_trad_rank.get("byRatePer100"),
                    "schoolCount": latest_trad_context.get("schoolCount"),
                    "medianRatePer100": latest_trad_context.get("medianRatePer100"),
                    "p95RatePer100": latest_trad_context.get("p95RatePer100"),
                },
                "sourceWorkbooks": sorted({row.get("sourceWorkbook") for row in rows_sorted if row.get("sourceWorkbook")}),
            }
        )

    rollup.sort(key=lambda row: (-(row.get("latestMetrics", {}).get("rawRestraintRatePer100") or -1), row.get("schoolName", "")))
    return rollup


def build_district_rollup(records: list[dict]) -> list[dict]:
    all_lens = build_district_year_metrics(records, use_traditional_lens=False)
    trad_lens = build_district_year_metrics(records, use_traditional_lens=True)
    apply_district_year_rankings(all_lens)
    apply_district_year_rankings(trad_lens)

    grouped: dict[str, list[dict]] = defaultdict(list)
    for row in records:
        if row.get("isSummaryRow"):
            continue
        grouped[row.get("districtCode", "")].append(row)

    rollup: list[dict] = []
    for district_code, rows in grouped.items():
        rows_sorted = sorted(rows, key=lambda row: row.get("schoolYear", ""))
        years = sorted({row.get("schoolYear") for row in rows_sorted if row.get("schoolYear")})

        yearly = []
        for year in years:
            all_metrics = all_lens.get(year, {}).get(district_code)
            if not all_metrics:
                continue
            trad_metrics = trad_lens.get(year, {}).get(district_code)
            yearly.append(
                {
                    **all_metrics,
                    "statewideAllRows": {
                        "districtCount": all_metrics.get("statewideContext", {}).get("districtCount"),
                        "medianRatePer100": all_metrics.get("statewideContext", {}).get("medianRatePer100"),
                        "p95RatePer100": all_metrics.get("statewideContext", {}).get("p95RatePer100"),
                        "rankByTotalRestraints": all_metrics.get("statewideRank", {}).get("byTotalRestraints"),
                        "rankByRatePer100": all_metrics.get("statewideRank", {}).get("byRatePer100"),
                    },
                    "statewideTraditional": {
                        "districtCount": trad_metrics.get("statewideContext", {}).get("districtCount") if trad_metrics else 0,
                        "medianRatePer100": trad_metrics.get("statewideContext", {}).get("medianRatePer100") if trad_metrics else None,
                        "p95RatePer100": trad_metrics.get("statewideContext", {}).get("p95RatePer100") if trad_metrics else None,
                        "rankByTotalRestraints": trad_metrics.get("statewideRank", {}).get("byTotalRestraints") if trad_metrics else None,
                        "rankByRatePer100": trad_metrics.get("statewideRank", {}).get("byRatePer100") if trad_metrics else None,
                    },
                }
            )

        visible_years = [item for item in yearly if item.get("rawRestraintRatePer100") is not None]
        latest = visible_years[-1] if visible_years else None
        previous = visible_years[-2] if len(visible_years) > 1 else None
        peak_total = max(visible_years, key=lambda item: item.get("rawTotals", {}).get("restraints") or 0, default=None)
        peak_rate = max(visible_years, key=lambda item: item.get("rawRestraintRatePer100") or 0.0, default=None)

        school_history: dict[str, list[dict]] = defaultdict(list)
        for row in rows_sorted:
            if row.get("isRawKnownRestraints"):
                school_history[row.get("schoolKey", "")].append(row)

        largest_jump = None
        for history in school_history.values():
            history_sorted = sorted(history, key=lambda row: row.get("schoolYear", ""))
            for idx in range(1, len(history_sorted)):
                current = history_sorted[idx]
                prior = history_sorted[idx - 1]
                delta = (current.get("rawTotalRestraints") or 0) - (prior.get("rawTotalRestraints") or 0)
                candidate = {
                    "schoolKey": current.get("schoolKey"),
                    "schoolCode": current.get("schoolCode"),
                    "schoolName": current.get("schoolName"),
                    "fromYear": prior.get("schoolYear"),
                    "toYear": current.get("schoolYear"),
                    "fromRestraints": prior.get("rawTotalRestraints"),
                    "toRestraints": current.get("rawTotalRestraints"),
                    "deltaRestraints": delta,
                }
                if largest_jump is None or delta > largest_jump.get("deltaRestraints", -10**9):
                    largest_jump = candidate

        latest_total = latest.get("rawTotals", {}).get("restraints") if latest else None
        top_school = latest.get("topSchool") if latest else None
        top_share = ratio((top_school or {}).get("rawTotalRestraints"), latest_total)

        rollup.append(
            {
                "districtCode": district_code,
                "districtName": rows_sorted[0].get("districtName", district_code),
                "years": years,
                "rowCount": len(rows_sorted),
                "visibleYearCount": len(visible_years),
                "latestYear": latest.get("schoolYear") if latest else None,
                "latestMetrics": {
                    "rawTotalRestraints": latest.get("rawTotals", {}).get("restraints") if latest else None,
                    "rawStudentsRestrained": latest.get("rawTotals", {}).get("studentsRestrained") if latest else None,
                    "rawTotalInjuries": latest.get("rawTotals", {}).get("injuries") if latest else None,
                    "rawRestraintRatePer100": latest.get("rawRestraintRatePer100") if latest else None,
                    "rawStudentsRestrainedRatePer100": latest.get("rawStudentsRestrainedRatePer100") if latest else None,
                    "rawInjuryRatePer100": latest.get("rawInjuryRatePer100") if latest else None,
                    "restraintsPerStudentRestrainedRaw": latest.get("restraintsPerStudentRestrainedRaw") if latest else None,
                    "injuriesPerRestraintRaw": latest.get("injuriesPerRestraintRaw") if latest else None,
                },
                "previousVisibleYear": previous.get("schoolYear") if previous else None,
                "yearOverYear": {
                    "deltaRestraints": (latest.get("rawTotals", {}).get("restraints") or 0) - (previous.get("rawTotals", {}).get("restraints") or 0)
                    if latest and previous
                    else None,
                    "deltaRatePer100": round((latest.get("rawRestraintRatePer100") or 0.0) - (previous.get("rawRestraintRatePer100") or 0.0), 4)
                    if latest and previous and latest.get("rawRestraintRatePer100") is not None and previous.get("rawRestraintRatePer100") is not None
                    else None,
                    "pctChangeRestraints": ratio(
                        ((latest.get("rawTotals", {}).get("restraints") or 0) - (previous.get("rawTotals", {}).get("restraints") or 0))
                        * 100.0,
                        previous.get("rawTotals", {}).get("restraints") if previous else None,
                    )
                    if latest and previous
                    else None,
                },
                "peakMetrics": {
                    "peakTotalYear": peak_total.get("schoolYear") if peak_total else None,
                    "peakTotalRestraints": peak_total.get("rawTotals", {}).get("restraints") if peak_total else None,
                    "peakRateYear": peak_rate.get("schoolYear") if peak_rate else None,
                    "peakRatePer100": peak_rate.get("rawRestraintRatePer100") if peak_rate else None,
                },
                "concentrationLatest": {
                    "topSchool": top_school,
                    "topSchoolShare": top_share,
                    "topSchoolSharePercent": round(top_share * 100.0, 2) if top_share is not None else None,
                },
                "largestSchoolJump": largest_jump,
                "statewideLatestAllRows": latest.get("statewideAllRows") if latest else {},
                "statewideLatestTraditional": latest.get("statewideTraditional") if latest else {},
                "yearly": yearly,
            }
        )

    rollup.sort(key=lambda row: (-(row.get("latestMetrics", {}).get("rawRestraintRatePer100") or -1), row.get("districtName", "")))
    return rollup


def build_statewide_totals(records: list[dict]) -> list[dict]:
    totals = [row for row in records if row.get("isSummaryRow")]
    totals.sort(key=lambda row: row.get("schoolYear", ""))
    return [
        {
            "schoolYear": row.get("schoolYear"),
            "enrollment": row.get("enrollment"),
            "rawTotalRestraints": row.get("rawTotalRestraints"),
            "rawStudentsRestrained": row.get("rawStudentsRestrained"),
            "rawTotalInjuries": row.get("rawTotalInjuries"),
            "rawRestraintRatePer100": row.get("rawRestraintRatePer100"),
            "rawStudentsRestrainedRatePer100": row.get("rawStudentsRestrainedRatePer100"),
            "rawInjuryRatePer100": row.get("rawInjuryRatePer100"),
            "sourceWorkbook": row.get("sourceWorkbook"),
        }
        for row in totals
    ]


def build_restraint_datasets(root: Path, workbooks: list[Path]) -> dict:
    all_school_year_rows = build_school_year_records(workbooks)
    add_school_year_rankings(all_school_year_rows)

    entity_rows = [row for row in all_school_year_rows if not row.get("isSummaryRow")]
    district_rollup = build_district_rollup(entity_rows)
    school_rollup = build_school_rollup(entity_rows)
    statewide_totals = build_statewide_totals(all_school_year_rows)

    return {
        "schoolYearRecords": all_school_year_rows,
        "districtRollup": district_rollup,
        "schoolRollup": school_rollup,
        "statewideTotals": statewide_totals,
    }


def run_restraint_pipeline(root: Path) -> dict:
    eml_result = extract_eml_attachments(root)
    loose_result = ingest_loose_intake_workbooks(root)
    workbooks = select_workbooks(root)
    datasets = build_restraint_datasets(root, workbooks)

    write_json(root / "data" / "restraint_school_year.json", {"records": datasets["schoolYearRecords"]})
    write_json(root / "data" / "restraint_district_rollup.json", {"records": datasets["districtRollup"]})
    write_json(root / "data" / "restraint_school_rollup.json", {"records": datasets["schoolRollup"]})
    write_json(root / "data" / "restraint_statewide_totals.json", {"records": datasets["statewideTotals"]})
    write_json(
        root / "data" / "restraint_explorer.json",
        {
            "districts": datasets["districtRollup"],
            "schools": datasets["schoolRollup"],
            "schoolYears": datasets["schoolYearRecords"],
            "statewideTotals": datasets["statewideTotals"],
        },
    )

    manifest = {
        "source": "DESE Student Restraints Report - Public Schools",
        "files": [rel_path(path, root) for path in workbooks],
        "yearsCovered": sorted({row.get("schoolYear") for row in datasets["schoolYearRecords"] if row.get("schoolYear")}),
        "suppressionRule": "Dash '-' indicates a DESE-suppressed value. Suppressed values are kept as unknown and not converted to midpoint estimates.",
        "rankingLenses": {
            "allRowsMinusSummary": "All DESE school rows with numeric restraint values, excluding only the Public Schools Total summary row.",
            "traditionalComparable": "Rows that pass both checks: enrollment >= 100 and school name does not include special-program keywords.",
        },
        "emlWorkbookAttachments": eml_result.get("workbookRecords", []),
        "looseIntakeWorkbooks": loose_result.get("workbookRecords", []),
        "referenceDocs": eml_result.get("referenceDocs", []),
    }
    write_json(root / "data" / "restraint_manifest.json", manifest)

    return {
        "workbookCount": len(workbooks),
        "schoolYearCount": len(datasets["schoolYearRecords"]),
        "districtCount": len(datasets["districtRollup"]),
        "schoolCount": len(datasets["schoolRollup"]),
    }
